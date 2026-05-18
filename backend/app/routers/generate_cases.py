import json
import csv
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from typing import Optional
from app.models.database import get_db
from app.models.models import TestCase, TestSuiteFunction, TestFunctionCase
from app.services.llm_service import (
    generate_test_cases_from_url,
    generate_test_cases_from_prd,
)
from app.utils import format_datetime

router = APIRouter(prefix="/api/generate", tags=["generation"])


def _parse_llm_cases(raw_cases: list) -> list:
    """标准化LLM返回的用例字段"""
    result = []
    for c in raw_cases:
        case_type = c.get("case_type", "positive")
        if case_type not in ("positive", "negative"):
            case_type = "positive"
        result.append({
            "name": c.get("name", c.get("用例名称", "未命名用例")),
            "case_type": case_type,
            "description": c.get("description", c.get("用例描述", "")),
            "focus_point": c.get("focus_point", c.get("关注点", "")),
            "preconditions": c.get("preconditions", c.get("前提条件", "")),
            "expected_result": c.get("expected_result", c.get("预期结果", "")),
        })
    return result


@router.post("/from-url")
async def generate_from_url(data: dict, db: Session = Depends(get_db)):
    """从URL生成测试用例"""
    url = data.get("url")
    if not url:
        raise HTTPException(status_code=400, detail="URL不能为空")

    function_id = data.get("function_id")
    description = data.get("description", "")

    try:
        cases_data = await generate_test_cases_from_url(db, url, description)
    except Exception as e:
        import traceback
        print(f"[ERROR] generate_from_url: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"AI生成失败: {str(e)}")

    cases_data = _parse_llm_cases(cases_data)

    # 如果没有function_id，只返回用例不入库
    if not function_id:
        return {
            "message": f"生成了{len(cases_data)}个测试用例（未关联功能）",
            "cases": cases_data,
        }

    created_cases = []
    for case_data in cases_data:
        tc = TestCase(
            function_id=function_id,
            name=case_data["name"],
            case_type=case_data["case_type"],
            description=case_data["description"],
            focus_point=case_data["focus_point"],
            preconditions=case_data["preconditions"],
            expected_result=case_data["expected_result"],
        )
        db.add(tc)
        db.flush()
        
        # 创建功能-用例关联
        assoc = TestFunctionCase(
            function_id=function_id,
            test_case_id=tc.id
        )
        db.add(assoc)
        created_cases.append(tc)

    db.commit()

    return {
        "message": f"生成了{len(created_cases)}个测试用例",
        "cases": [
            {"id": tc.id, "name": tc.name, "case_type": tc.case_type}
            for tc in created_cases
        ],
    }


@router.post("/from-prd")
async def generate_from_prd(data: dict, db: Session = Depends(get_db)):
    """从需求文档文本内容生成测试用例"""
    prd_content = data.get("content")
    if not prd_content:
        raise HTTPException(status_code=400, detail="需求文档内容不能为空")

    function_id = data.get("function_id")

    try:
        cases_data = await generate_test_cases_from_prd(db, prd_content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI生成失败: {str(e)}")

    cases_data = _parse_llm_cases(cases_data)

    created_cases = []
    for case_data in cases_data:
        tc = TestCase(
            function_id=function_id,
            name=case_data["name"],
            case_type=case_data["case_type"],
            description=case_data["description"],
            focus_point=case_data["focus_point"],
            preconditions=case_data["preconditions"],
            expected_result=case_data["expected_result"],
        )
        db.add(tc)
        db.flush()
        
        # 创建功能-用例关联
        assoc = TestFunctionCase(
            function_id=function_id,
            test_case_id=tc.id
        )
        db.add(assoc)
        created_cases.append(tc)

    db.commit()

    return {
        "message": f"生成了{len(created_cases)}个测试用例",
        "cases": [
            {"id": tc.id, "name": tc.name, "case_type": tc.case_type}
            for tc in created_cases
        ],
    }


@router.post("/upload-prd")
async def upload_prd(
    function_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """上传PRD文档（txt/doc/docx）生成测试用例"""
    filename = file.filename.lower()
    content_bytes = await file.read()

    # 解析文档内容
    if filename.endswith(".txt"):
        text = content_bytes.decode("utf-8", errors="ignore")
    elif filename.endswith(".docx") or filename.endswith(".doc"):
        try:
            from docx import Document
            doc = Document(io.BytesIO(content_bytes))
            text = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Word文档解析失败: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="仅支持 txt、doc、docx 格式")

    if not text.strip():
        raise HTTPException(status_code=400, detail="文档内容为空")

    # 调用LLM生成用例
    try:
        cases_data = await generate_test_cases_from_prd(db, text[:8000])
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI生成失败: {str(e)}")

    cases_data = _parse_llm_cases(cases_data)

    created_cases = []
    for case_data in cases_data:
        tc = TestCase(
            function_id=function_id,
            name=case_data["name"],
            case_type=case_data["case_type"],
            description=case_data["description"],
            focus_point=case_data["focus_point"],
            preconditions=case_data["preconditions"],
            expected_result=case_data["expected_result"],
        )
        db.add(tc)
        db.flush()
        created_cases.append(tc)

    db.commit()

    return {
        "message": f"从文档生成了{len(created_cases)}个测试用例",
        "cases": [
            {"id": tc.id, "name": tc.name, "case_type": tc.case_type}
            for tc in created_cases
        ],
    }


@router.post("/import-excel")
async def import_excel(
    function_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """从Excel导入测试用例

    Excel表头应包含以下列（中文或英文均可）：
    用例名称/name, 类型/case_type, 描述/description,
    关注点/focus_point, 前提条件/preconditions, 预期结果/expected_result
    """
    filename = file.filename.lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="仅支持 xlsx/xls 格式")

    content_bytes = await file.read()

    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content_bytes), read_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel解析失败: {str(e)}")

    # 读取表头
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel为空")

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    # 字段映射
    field_map = {
        "用例名称": "name", "name": "name", "名称": "name",
        "类型": "case_type", "case_type": "case_type", "type": "case_type",
        "描述": "description", "description": "description", "用例描述": "description",
        "关注点": "focus_point", "focus_point": "focus_point",
        "前提条件": "preconditions", "preconditions": "preconditions",
        "预期结果": "expected_result", "expected_result": "expected_result",
    }

    col_indices = {}
    for i, h in enumerate(headers):
        mapped = field_map.get(h)
        if mapped:
            col_indices[mapped] = i

    if "name" not in col_indices:
        raise HTTPException(status_code=400, detail="Excel必须包含'用例名称'或'name'列")

    created_cases = []
    for row in rows[1:]:
        if not row or not row[col_indices["name"]]:
            continue

        name = str(row[col_indices["name"]]).strip()
        case_type_raw = str(row[col_indices.get("case_type", -1)]).strip().lower() if "case_type" in col_indices else "positive"
        if case_type_raw in ("反例", "negative", "neg"):
            case_type = "negative"
        else:
            case_type = "positive"

        tc = TestCase(
            function_id=function_id,
            name=name,
            case_type=case_type,
            description=str(row[col_indices.get("description", -1)]).strip() if "description" in col_indices and row[col_indices["description"]] else "",
            focus_point=str(row[col_indices.get("focus_point", -1)]).strip() if "focus_point" in col_indices and row[col_indices["focus_point"]] else "",
            preconditions=str(row[col_indices.get("preconditions", -1)]).strip() if "preconditions" in col_indices and row[col_indices["preconditions"]] else "",
            expected_result=str(row[col_indices.get("expected_result", -1)]).strip() if "expected_result" in col_indices and row[col_indices["expected_result"]] else "",
        )
        db.add(tc)
        db.flush()
        created_cases.append(tc)

    db.commit()
    wb.close()

    return {
        "message": f"导入了{len(created_cases)}个测试用例",
        "cases": [
            {"id": tc.id, "name": tc.name, "case_type": tc.case_type}
            for tc in created_cases
        ],
    }


@router.post("/import")
async def import_cases(
    function_id: int = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """导入测试用例（支持JSON/CSV，兼容旧接口）"""
    content = await file.read()
    filename = file.filename.lower()

    if filename.endswith(".json"):
        try:
            cases_data = json.loads(content)
            if isinstance(cases_data, dict):
                cases_data = cases_data.get("cases", [cases_data])
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="JSON格式错误")
    elif filename.endswith(".csv"):
        try:
            reader = csv.DictReader(io.StringIO(content.decode("utf-8")))
            cases_data = list(reader)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"CSV解析错误: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="仅支持JSON和CSV格式")

    cases_data = _parse_llm_cases(cases_data)

    created_cases = []
    for case_data in cases_data:
        tc = TestCase(
            function_id=function_id,
            name=case_data["name"],
            case_type=case_data["case_type"],
            description=case_data["description"],
            focus_point=case_data["focus_point"],
            preconditions=case_data["preconditions"],
            expected_result=case_data["expected_result"],
        )
        db.add(tc)
        db.flush()
        created_cases.append(tc)

    db.commit()

    return {
        "message": f"导入了{len(created_cases)}个测试用例",
        "cases": [
            {"id": tc.id, "name": tc.name, "case_type": tc.case_type}
            for tc in created_cases
        ],
    }


@router.post("/functions/{function_id}/generate-from-description")
async def generate_cases_from_description(
    function_id: int,
    description: str = Form(...),
    db: Session = Depends(get_db)
):
    """从自然语言描述生成测试用例"""
    from app.models.models import TestFunction
    from app.services.llm_service import generate_cases_from_text
    
    # 验证功能存在
    function = db.query(TestFunction).filter(TestFunction.id == function_id).first()
    if not function:
        raise HTTPException(status_code=404, detail="功能不存在")
    
    try:
        cases_data = await generate_cases_from_text(db, 
            description=description,
            function_name=function.name
        )
        
        created_cases = []
        for case_data in cases_data:
            new_case = TestCase(function_id=function_id, 
                name=case_data.get('name', ''),
                case_type=case_data.get('type', '正例'),
                description=case_data.get('description', ''),
                focus_point=case_data.get('focus_point', ''),
                preconditions=case_data.get('preconditions', ''),
                expected_result=case_data.get('expected_result', ''),
                status='未测试'
            )
            db.add(new_case)
            db.flush()
            
            # 关联到功能
            association = TestFunctionCase(
                function_id=function_id,
                test_case_id=new_case.id
            )
            db.add(association)
            created_cases.append(new_case)
        
        db.commit()
        
        return {
            "message": f"成功生成 {len(created_cases)} 个测试用例",
            "count": len(created_cases),
            "cases": [{"id": c.id, "name": c.name} for c in created_cases]
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
